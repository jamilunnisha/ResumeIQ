from openpyxl import Workbook
from pathlib import Path


def export_to_excel(data: dict, filename: str):

    output_dir = Path(__file__).resolve().parent.parent / "database"
    output_dir.mkdir(exist_ok=True)

    excel_path = output_dir / "parsed_resumes.xlsx"

    if excel_path.exists():
        from openpyxl import load_workbook
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active

        sheet.title = "Candidates"

        sheet.append([
            "Name",
            "Email",
            "Phone",
            "Experience",
            "Skills"
        ])

    skills = ", ".join(data.get("skills", []))

    sheet.append([
        data.get("name"),
        data.get("email"),
        data.get("phone"),
        data.get("experience"),
        skills
    ])

    workbook.save(excel_path)

    return str(excel_path)